from argparse import ArgumentParser, BooleanOptionalAction
import json
from pathlib import Path
import random

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment


def read_files(files: list[str], data_field: str = None) -> dict:
    """Reads the input JSON files and returns all the data in a dictionary.

    Args:
        files (list[str]): List of input files.
        data_field (str, optional): If the file is a JSON, the data might be in a specific field. Defaults to None.

    Raises:
        ValueError: All the files must be JSON.

    Returns:
        dict: Dictionary where the keys are the filenames and the values are lists of dialogues.
    """

    data = {}

    for filename in files:
        filepath = Path(filename)

        # Check if the file is a JSON
        if not filepath.is_file() or filepath.suffix != ".json":
            raise ValueError(f"{filename} must be a JSON file")

        # Read the file
        with open(filename, "r") as f:
            file_data = json.load(f)
            data[filename] = file_data[data_field] if data_field else file_data

    return data


def get_response(sample: dict) -> str:
    """Given a sample, returns the appropriate response.
    Works with results from retrieval or from genration

    Args:
        sample (dict): A sample from retrieval or generation.

    Raises:
        ValueError: If the sample does not contain a model_answer or knowledge.

    Returns:
        str: The response.
    """
    if "model_answer" in sample:
        return sample["model_answer"]
    elif "knowledge" in sample:
        return sample["knowledge"][0]
    else:
        raise ValueError("The sample does not contain a model_answer or knowledge")


def format_samples(data: dict, ground_truth: bool = True) -> list[dict]:
    """Merges each sample of all datasets into a single dict, returning a list with all samples.

    Args:
        data (dict): Dictionary where the keys are the filenames and the values are lists of dialogues.
        ground_truth (bool, optional): Wheather to include the ground truth resposne in the answers. Defaults to True.

    Raises:
        ValueError: If the datasets are misaligned (might have different samples).

    Returns:
        list[dict]: List where each sample consists of a id, context and responses (ground truth and models).
    """

    dataset = []
    names = data.keys()

    for samples in zip(*data.values()):
        if len(set(sample["id"] for sample in samples)) > 1:
            raise ValueError("Mismatch between samples ids.")

        dataset.append(
            {
                "id": samples[0]["id"],
                "context": samples[0]["context"],
                "response": (
                    {"ground_truth": samples[0]["delexicalized"]}
                    if ground_truth
                    else {}
                )
                | {name: get_response(sample) for name, sample in zip(names, samples)},
            }
        )

    return dataset


def shuffle_responses(data: list[dict], seed: int = 42) -> list[dict]:
    """Receives the dataset and shuffles the order of the responses, for an unbias human evaluation.

    Args:
        data (list[dict]): List where each sample consists of a id, context and responses (ground truth and models).
        seed (int, optional): Seed for random numbers. Defaults to 42.

    Returns:
        list[dict]: Original list but with responses shuffled. The field "response" is now a dict of index -> response str.
                    The new field "shuffle_order" is a dict with the origin of the response and its index in "response".
    """
    random.seed(seed)

    for sample in data:
        responses = list(sample["response"].items())
        random.shuffle(responses)
        sample["response"] = {i: x[1] for i, x in enumerate(responses)}
        sample["shuffle_order"] = {x[0]: i for i, x in enumerate(responses)}

    return data


def format_excel(filename: str) -> None:
    """Formats the row and columns sizes of the excel file. Also aligns the cells

    Args:
        filename (str): Path to the .xlsx file to save the data on.

    Raises:
        ValueError: _description_
    """

    filepath = Path(filename)

    # Format the rows height and column width
    wb = load_workbook(filepath)
    ws = wb.active

    # Get index of desired column
    for column_cell in ws.iter_cols(1, ws.max_column):
        if column_cell[0].value == "context":
            column_index = column_cell[0].column
            break
    else:
        raise ValueError(
            f'The file {filepath} does not contain a columns named "context".'
        )

    # Set the row height
    for row in range(2, ws.max_row + 1):
        # Calculate the number of lines in the cell value
        cell = ws.cell(row=row, column=column_index)
        num_lines = str(cell.value).count("\n") + 1

        # Set the row height based on the number of lines
        if num_lines > 1:
            row_dimension = ws.row_dimensions[row]
            row_dimension.height = 14 * num_lines

    # Set the column width
    for col in ws.columns:
        col_letter = col[0].column_letter
        col_name = str(col[0].value)

        # Width for each column
        if col_name == "id":
            width = 15
            horizontal = "center"
            vertical = "center"
        elif col_name == "context":
            width = 100
            horizontal = "left"
            vertical = "top"
        elif col_name.startswith("best"):
            width = 3
            horizontal = "center"
            vertical = "center"
        else:
            width = 100
            horizontal = "left"
            vertical = "bottom"

        # Set dimensions
        ws.column_dimensions[col_letter].width = width

        # Set alignment
        alignment = Alignment(horizontal=horizontal, vertical=vertical, wrap_text=True)
        for cell in col:
            cell.alignment = alignment

    # Freeze the top row and first two columns
    ws.freeze_panes = ws.cell(row=2, column=3)

    # Save the workbook
    wb.save(filepath)
    print(f"Saved formatted excel to: {filepath}")


def save_to_xlsx(data: list[dict], n_samples: int, filename: str) -> None:
    """Saves the data and the order of the data in two different files.
    Such that an evaluator can see the data without knowing the order.
    The order is saved in an additional file in the same directory as filename but with the suffix "_order".
    It randomly selects n_samples.

    Args:
        data (list[dict]): List of samples where the responses are shuffled.
        int (int): Number of samples to save
        filename (str): Path to the .xlsx file to save the data on.
    """
    # Select random samples
    data = random.sample(data, n_samples)

    # Format
    formatted = [
        {
            "id": sample["id"],
            "context": "\n".join(sample["context"]),
        }
        | {
            i + 1: response.lstrip("System: ")
            for i, response in sample["response"].items()
        }
        | {f"best {i}": "" for i in range(1, len(sample["response"]) + 1)}
        for sample in data
    ]

    order = [list(sample["shuffle_order"].keys()) for sample in data]

    # Save to excel
    filepath = Path(filename)
    order_filepath = filepath.with_stem(filepath.stem + "_order")

    pd.DataFrame(formatted).to_excel(filepath, index=False)
    print(f"Saved data to: {filepath}")
    pd.DataFrame(order).to_excel(order_filepath, index=False, header=False)
    print(f"Saved secret order to: {order_filepath}")

    # Improve formatting of file
    format_excel(filename)


if __name__ == "__main__":
    parser = ArgumentParser()
    parser.add_argument(
        "-i",
        "--input",
        nargs="+",
        type=str,
        required=True,
        help="Files with text to be evaluated.",
    )
    parser.add_argument(
        "--ground_truth",
        default=False,
        action=BooleanOptionalAction,
        help="Include ground truth responses.",
    )
    parser.add_argument(
        "--data_field",
        type=str,
        default=None,
        help="If the input file is a JSON, the data might be in a particular field.",
    )
    parser.add_argument(
        "-seed", type=str, default=42, help="Seed to initialize randomness."
    )
    parser.add_argument(
        "--n_samples",
        type=int,
        default=100,
        help="Number of samples to save in .xlsx file.",
    )
    parser.add_argument(
        "--output", type=str, help="Filepath to save the results in an .xslx file."
    )
    args = parser.parse_args()

    data = read_files(args.input, args.data_field)
    data = format_samples(data, args.ground_truth)
    data = shuffle_responses(data, args.seed)

    if args.output:
        save_to_xlsx(data, args.n_samples, args.output)
    else:
        print(json.dumps(data, indent=4))
